"""
experiment_L.py — Hyperparameter sweep over OrthLayer Newton-Schulz iteration count L.

Sweeps L from 0 to 10 (3 runs each), evaluates every value on the NUDT-SIRST
test set, and writes a summary Excel file.

Usage:
    python experiment_L.py [--epochs 400] [--batchSize 4] [--threads 1]

Architecture follows ``train.py`` — unified Net wrapper, SoftIoULoss,
MultiStepLR scheduler, and full metric evaluation (mIoU, PD, FA, F-score).
"""

import argparse
import os
import sys
import time
import numpy as np
import torch
from torch.autograd import Variable
from torch.utils.data import DataLoader

from net import Net
from utils.utils import seed_pytorch, get_optimizer
from utils.datasets import NUDTSIRSTSetLoader
from evaluation.mIoU import mIoU
from evaluation.pd_fa import PD_FA
from evaluation.TPFNFP import SegmentationMetricTPFNFP

# ---------------------------------------------------------------------------
#  CLI
# ---------------------------------------------------------------------------
parser = argparse.ArgumentParser(description="Sweep OrthLayer L (iteration count)")

parser.add_argument("--epochs", type=int, default=400,
                    help="Training epochs per run (default: 400)")
parser.add_argument("--batchSize", type=int, default=4)
parser.add_argument("--threads", type=int, default=1,
                    help="DataLoader workers")
parser.add_argument("--lr", type=float, default=5e-4,
                    help="Learning rate")
parser.add_argument("--save_dir", default="./experiment_logs/L_sweep",
                    help="Root directory for checkpoints and logs")
parser.add_argument("--dataset_dir", default="./data/NUDT-SIRST/",
                    help="NUDT-SIRST dataset root")

opt = parser.parse_args()

# ---------------------------------------------------------------------------
#  Derived settings (mirrors train.py defaults)
# ---------------------------------------------------------------------------
L_VALUES = list(range(0, 11))          # 0 … 10
N_RUNS = 3
BASE_SEED = 42
OPTIMIZER_NAME = "Adam"
SCHEDULER_NAME = "MultiStepLR"
SCHEDULER_STEP = [opt.epochs // 2, int(opt.epochs * 0.75)]   # e.g. [200, 300] for 400

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
torch.cuda.set_device(0)

# Ensure save root exists
os.makedirs(opt.save_dir, exist_ok=True)

# ---------------------------------------------------------------------------
#  Helpers (mirrors train.py)
# ---------------------------------------------------------------------------
def save_checkpoint(state, save_path):
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    torch.save(state, save_path)


def evaluate(net, test_loader, threshold=0.5):
    """Run full metric suite on the test set.  Returns (mIoU, PD, FA, F-score)."""
    eval_miou = mIoU()
    eval_pd_fa = PD_FA()
    eval_f = SegmentationMetricTPFNFP(nclass=1)

    for img, gt_mask, size, _ in test_loader:
        with torch.no_grad():
            img = Variable(img).cuda()
            pred = net.forward(img, mode="test")
            pred = pred[:, :, : size[0], : size[1]]
        gt_mask = gt_mask[:, :, : size[0], : size[1]]

        eval_miou.update((pred > threshold).cpu(), gt_mask)
        eval_pd_fa.update(
            pred[0, 0].cpu().detach().numpy(),
            gt_mask[0, 0].detach().numpy(),
            size,
        )
        eval_f.update(
            labels=gt_mask[0, 0].detach().numpy(),
            preds=pred[0, 0].cpu().detach().numpy(),
        )

    _, Yin_mIoU = eval_miou.get()
    pd, fa = eval_pd_fa.get()
    _, _, _, fscore = eval_f.get()
    return Yin_mIoU, pd, fa, fscore


def train_one_run(L_value, run_idx, seed):
    """
    Train SFBD_Net with a specific L on NUDT-SIRST for ``opt.epochs`` epochs.
    Returns the best metrics seen (after epoch > epochs/2).
    """
    dataset_dir = opt.dataset_dir
    train_set = NUDTSIRSTSetLoader(base_dir=dataset_dir, mode="trainval")
    test_set = NUDTSIRSTSetLoader(base_dir=dataset_dir, mode="test")

    train_loader = DataLoader(
        dataset=train_set, num_workers=opt.threads,
        batch_size=opt.batchSize, shuffle=True,
    )
    test_loader = DataLoader(
        dataset=test_set, num_workers=1, batch_size=1, shuffle=False,
    )

    # ---- Model ----
    net = Net(model_name="SFBD_Net", mode="train", L=L_value).cuda()
    net.train()

    # ---- Optimiser & scheduler ----
    optimizer_settings = {"lr": opt.lr}
    scheduler_settings = {
        "epochs": opt.epochs,
        "step": SCHEDULER_STEP,
        "gamma": 0.1,
    }
    optimizer, scheduler = get_optimizer(
        net, OPTIMIZER_NAME, SCHEDULER_NAME,
        optimizer_settings, scheduler_settings,
    )

    # ---- Bookkeeping ----
    total_loss_epoch = []
    best_miou = 0.0
    best_fscore = 0.0
    best_pd = 0.0
    best_fa = 1.0

    # ---- Training loop ----
    for idx_epoch in range(opt.epochs):
        for img, gt_mask in train_loader:
            img, gt_mask = Variable(img).cuda(), Variable(gt_mask).cuda()
            if img.shape[0] == 1:
                continue
            pred = net.forward(img, mode="train")
            loss = net.loss(pred, gt_mask, img)
            total_loss_epoch.append(loss.detach().cpu())

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

        scheduler.step()

        # ---- Evaluate after half the epochs ----
        if idx_epoch + 1 > opt.epochs // 2:
            avg_loss = float(np.array(total_loss_epoch).mean())
            total_loss_epoch = []

            net.eval()
            miou, pd, fa, fscore = evaluate(net, test_loader)
            net.train()

            if miou > best_miou:
                best_miou = miou
            if fscore > best_fscore:
                best_fscore = fscore
            if pd > best_pd:
                best_pd = pd
            if fa < best_fa:
                best_fa = fa

            print(f"    [L={L_value} run={run_idx}] epoch {idx_epoch+1:3d}  "
                  f"loss={avg_loss:.4f}  mIoU={miou:.4f}  fscore={fscore:.4f}")

    return {
        "L": L_value,
        "run": run_idx,
        "seed": seed,
        "best_mIoU": best_miou,
        "best_PD": best_pd,
        "best_FA": best_fa,
        "best_Fscore": best_fscore,
    }


# ---------------------------------------------------------------------------
#  Main sweep
# ---------------------------------------------------------------------------
def main():
    all_results = []

    print(f"{'='*60}")
    print(f"L sweep: {L_VALUES}  ×  {N_RUNS} runs  ×  {opt.epochs} epochs")
    print(f"Save dir: {opt.save_dir}")
    print(f"{'='*60}")

    for L_val in L_VALUES:
        for run in range(1, N_RUNS + 1):
            seed = BASE_SEED + run * 100 + L_val
            seed_pytorch(seed)

            run_name = f"L{L_val:02d}_run{run}"
            print(f"\n>>> {run_name}  (seed={seed})")

            t0 = time.time()
            result = train_one_run(L_val, run, seed)
            elapsed = time.time() - t0

            print(f"<<< {run_name}  mIoU={result['best_mIoU']:.4f}  "
                  f"Fscore={result['best_Fscore']:.4f}  [{elapsed:.0f}s]")
            all_results.append(result)

    # ---- Export Excel ----
    _write_excel(all_results)


def _write_excel(results):
    """Write sweep results to an .xlsx file."""
    try:
        import openpyxl
    except ImportError:
        print("\n[WARN] openpyxl not installed — writing CSV instead.")
        _write_csv(results)
        return

    wb = openpyxl.Workbook()

    # ---- Sheet 1: raw results ----
    ws_raw = wb.active
    ws_raw.title = "raw"
    headers = ["L", "run", "seed", "best_mIoU", "best_PD", "best_FA", "best_Fscore"]
    ws_raw.append(headers)
    for r in results:
        ws_raw.append([r[h] for h in headers])

    # ---- Sheet 2: aggregated (mean ± std per L) ----
    ws_agg = wb.create_sheet("aggregated")
    agg_headers = ["L", "mean_mIoU", "std_mIoU", "mean_PD", "std_PD",
                   "mean_FA", "std_FA", "mean_Fscore", "std_Fscore"]
    ws_agg.append(agg_headers)
    for L_val in L_VALUES:
        group = [r for r in results if r["L"] == L_val]
        row = [L_val]
        for metric in ["best_mIoU", "best_PD", "best_FA", "best_Fscore"]:
            vals = [g[metric] for g in group]
            row.append(np.mean(vals))
            row.append(np.std(vals))
        ws_agg.append(row)

    # ---- Sheet 3: best per L ----
    ws_best = wb.create_sheet("best_per_L")
    ws_best.append(["L", "best_mIoU", "best_PD", "best_FA", "best_Fscore", "run"])
    for L_val in L_VALUES:
        group = [r for r in results if r["L"] == L_val]
        best = max(group, key=lambda x: x["best_mIoU"])
        ws_best.append([L_val, best["best_mIoU"], best["best_PD"],
                        best["best_FA"], best["best_Fscore"], best["run"]])

    path = os.path.join(opt.save_dir, "L_sweep_results.xlsx")
    wb.save(path)
    print(f"\nResults saved to {path}")


def _write_csv(results):
    """Fallback CSV output."""
    path = os.path.join(opt.save_dir, "L_sweep_results.csv")
    with open(path, "w") as f:
        headers = ["L", "run", "seed", "best_mIoU", "best_PD", "best_FA", "best_Fscore"]
        f.write(",".join(headers) + "\n")
        for r in results:
            f.write(",".join(str(r[h]) for h in headers) + "\n")
    print(f"\nResults saved to {path}")


if __name__ == "__main__":
    main()
